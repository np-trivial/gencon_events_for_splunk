import urllib.request
from urllib.error import HTTPError
import http.client
import json
import zipfile
import io
import pathlib
import re
import sys
from datetime import datetime, timezone
import csv
from openpyxl import load_workbook

eventzip = zipfile.ZipFile( sys.argv[1] )
eventxlsx = io.BytesIO(eventzip.read('events.xlsx'))

lmt = eventzip.getinfo('events.xlsx').date_time
last_modified = datetime(lmt[0], lmt[1], lmt[2], lmt[3], lmt[4], lmt[5], 0, tzinfo=timezone.utc).isoformat()

wb = load_workbook(eventxlsx)
ws = wb.active
# "clean keys", that is, change odd characters to underscore
header_row = [ 'File_Time' ] + [  re.sub("[^\tA-Za-z]","_",x.value) for x in ws[1] ]
csvout = csv.writer(sys.stdout, delimiter='\t', quoting=csv.QUOTE_ALL)
csvout.writerow(header_row)
all_rows = ws.rows
all_rows.__next__()     #skip the first row, it's the header
for i in all_rows:
    rowout = [ last_modified ] + [str(x.value).replace("\n","  ;;  ") for x in i]
    csvout.writerow(rowout)
