import urllib.request
from urllib.error import HTTPError
import http.client
import json
import zipfile
import io
import pathlib
import re
from openpyxl import load_workbook
import csv 
import sys

# a place to store the latest timestamp we've downloaded

last_modified_filename=str(pathlib.Path(__file__).resolve().parent) + '/../local/last_modified.log'

try:
  last_modified_fh=open(last_modified_filename, 'r')
  old_last_modified=json.loads(last_modified_fh.read())
  last_modified_fh.close()
except(Exception):
    # in case the timestamp file is empty (e.g. first run)
    old_last_modified=""

my_request = urllib.request.Request('https://www.gencon.com/downloads/events.zip')

# we get an HTTP 403 unless we include a user agent
my_request.add_header('User-agent', 'Mozilla/5.0')

# only add the If-Modified-Since header if we have it
if len(old_last_modified) > 0:
    my_request.add_header('If-Modified-Since',old_last_modified)    

try:
    with urllib.request.urlopen(my_request) as f:
        headers=f.info()

        last_modified=headers['Last-Modified']
        last_modified_fh=open(last_modified_filename, 'w')
        last_modified_fh.write(json.dumps(last_modified))
        last_modified_fh.close()

        #print(last_modified)

        eventzip = zipfile.ZipFile( io.BytesIO( f.read() ))
        eventxlsx = io.BytesIO(eventzip.read('events.xlsx'))

        wb = load_workbook(eventxlsx)
        ws = wb.active
        header_row = [ 'File_Time' ] + [  x.value for x in ws[1] ]
        csvout = csv.writer(sys.stdout, delimiter='\t', quoting=csv.QUOTE_ALL)
        csvout.writerow(header_row)
        all_rows = ws.rows
        all_rows.__next__()     #skip the first row, it's the header
        for i in ws.rows:
            rowout = [ last_modified ] + [str(x.value).replace("\n","  ;;  ") for x in i]
            csvout.writerow(rowout)

except HTTPError:
    # any kind of error (which includes "304 Not Modified")
    pass
